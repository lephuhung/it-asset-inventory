"""Test báo cáo Excel — sinh file hợp lệ, mask SĐT, filter theo RBAC."""
from __future__ import annotations

import io
import uuid

from openpyxl import load_workbook

from app.db.models import Machine, MachineSpec, UserRole


async def _seed_machine_with_user(
    session_factory,
    org_id,
    *,
    hostname="PC-EXPORT-01",
    status="online",
    phone="0987654321",
    email="export@test.gov.vn",
):
    """Tạo org owner + user + máy có spec (dùng cho test export)."""
    from app.core.security import encrypt_aes_gcm, hash_password
    from app.db.models import User

    async with session_factory() as s:
        admin = User(
            org_id=org_id,
            full_name="Quản trị export",
            email="exp_admin@test.gov.vn",
            role=UserRole.ADMIN_GLOBAL.value,
            password_hash=hash_password("x"),
        )
        s.add(admin)
        await s.flush()

        user = User(
            org_id=org_id,
            full_name="Nguyễn Văn Export",
            email=email,
            phone_encrypted=encrypt_aes_gcm(phone),
            role=UserRole.VIEWER.value,
        )
        s.add(user)
        await s.flush()

        m = Machine(
            org_id=org_id,
            machine_uuid="UUID-EXPORT-001",
            hostname=hostname,
            fingerprint={"smbios_uuid": "U-1"},
            status=status,
            assigned_user_id=user.id,
        )
        s.add(m)
        await s.flush()
        s.add(
            MachineSpec(
                machine_id=m.id,
                os_name="Windows 11",
                os_build="22631",
                cpu={"model": "Intel i7", "cores": 8},
                ram_gb=32.0,
                disks=[{"model": "NVMe", "capacity_gb": 512}],
            )
        )
        await s.commit()
        mid = m.id
    return mid


async def test_export_creates_valid_xlsx(client, seeded_env, session_factory):
    admin_token = await _login(client, seeded_env["email"], seeded_env["password"])
    org_id = seeded_env["org_id"]
    await _seed_machine_with_user(session_factory, uuid.UUID(org_id))

    r = await client.post(
        "/api/reports/export",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 200, r.text
    assert "application/vnd.openxmlformats" in r.headers["content-type"]
    assert ".xlsx" in r.headers["content-disposition"]

    # Parse Excel và kiểm tra nội dung
    wb = load_workbook(io.BytesIO(r.content))
    assert "Máy tính" in wb.sheetnames
    assert "Thống kê" in wb.sheetnames

    ws = wb["Máy tính"]
    headers = [c.value for c in ws[1]]
    assert "Hostname" in headers and "Số điện thoại" in headers and "Trạng thái" in headers

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    row = rows[0]
    assert "PC-EXPORT-01" in row  # hostname
    # Máy online + spec đầy đủ
    assert "Windows 11" in row
    assert 32.0 in row

    # SĐT phải được MASK mặc định theo format 0987•••321 (giữ đầu+cuối, giấu giữa)
    phone_idx = headers.index("Số điện thoại")
    phone_val = str(row[phone_idx])
    assert phone_val == "0987•••321", f"SĐT phải mask đúng format: {phone_val}"

    # Người dùng + email
    name_idx = headers.index("Người dùng")
    assert row[name_idx] == "Nguyễn Văn Export"


async def test_export_unauthorized(client):
    r = await client.post("/api/reports/export")
    assert r.status_code == 401


async def test_export_invalid_status(client, seeded_env):
    token = await _login(client, seeded_env["email"], seeded_env["password"])
    r = await client.post(
        "/api/reports/export?status=bogus",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert r.status_code == 400


async def test_export_with_phone_full_permission(client, seeded_env, session_factory):
    """Admin có quyền → include_phone_full=true hiện SĐT đầy đủ."""
    token = await _login(client, seeded_env["email"], seeded_env["password"])
    org_id = seeded_env["org_id"]
    await _seed_machine_with_user(session_factory, uuid.UUID(org_id), phone="0987654321")

    r = await client.post(
        "/api/reports/export?include_phone_full=true",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["Máy tính"]
    headers = [c.value for c in ws[1]]
    phone_idx = headers.index("Số điện thoại")
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert rows and rows[0][phone_idx] == "0987654321"


async def _login(client, email, password):
    r = await client.post("/api/auth/login", json={"email": email, "password": password})
    assert r.status_code == 200, r.text
    return r.json()["access_token"]