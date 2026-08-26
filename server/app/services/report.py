"""Báo cáo Excel — sinh workbook danh sách máy & thống kê (mục 5.4, Sprint 4).

Dùng openpyxl. Số điện thoại **mask mặc định**; `show_full_phone=True` chỉ khi
người gọi có quyền xem đầy đủ (RBAC — để dành Phase sau khi có phân quyền chi tiết).

Workbook gồm:
  • Sheet "Máy tính"       — danh sách máy chi tiết
  • Sheet "Thống kê"       — tổng quan theo status + theo tổ chức
  • Sheet "Thông tin thu thập" — thông báo tuân thủ, minh bạch dữ liệu (mục 7.4)
"""
from __future__ import annotations

import io
import uuid
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.phone_encryption import mask_phone

# ── Styles ─────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
WRAP = Alignment(wrap_text=True, vertical="top")


def _style_header(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def _auto_col_width(ws, ncols: int, max_row: int, min_width: int = 10, max_width: int = 45) -> None:
    for col in range(1, ncols + 1):
        width = min_width
        for row in range(1, max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                # Unicode width ~1.6 cho ký tự Đông Á/VN
                length = sum(1.6 if ord(ch) > 0x2E7F else 1.0 for ch in str(val))
                width = max(width, min(length + 2, max_width))
        ws.column_dimensions[get_column_letter(col)].width = width


def _fmt_dt(dt: datetime | None) -> str:
    if dt is None:
        return ""
    t = dt
    if t.tzinfo is not None:
        t = t.astimezone(UTC).replace(tzinfo=None)
    return t.strftime("%Y-%m-%d %H:%M")


async def build_machines_workbook(
    machines: list[Any],
    *,
    token_meta: dict[uuid.UUID, dict] | None = None,
    show_full_phone: bool = False,
    generated_by: str | None = None,
) -> bytes:
    """Tạo workbook Excel từ danh sách máy (ORM models). Trả về nội dung file bytes.

    `token_meta`: map machine_id → {full_name, email, department, position} lấy từ
    enroll_tokens (server gán khi enroll) — bổ sung khi user chưa có tài khoản users.
    """
    token_meta = token_meta or {}
    wb = Workbook()

    # ── Sheet 1: Máy tính ──────────────────────────────────────
    ws = wb.active
    ws.title = "Máy tính"
    headers = [
        "STT", "Hostname", "Mã máy (hash)", "Trạng thái", "Vòng đời", "Máy ảo?",
        "OS", "Build", "CPU", "RAM (GB)", "Ổ đĩa",
        "Người dùng", "Email", "Số điện thoại", "Phòng ban",
        "Tổ chức", "Lần cuối online", "Enrolled at", "Ghi chú",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for i, m in enumerate(machines, start=1):
        spec = _latest_spec(m)
        assigned = m.assigned_user
        phone_plain = None
        if m.assigned_user_id and hasattr(assigned, "phone_encrypted"):
            phone_plain = mask_phone(assigned.phone_encrypted) if not show_full_phone else _decrypt_phone(assigned)
        tmeta = token_meta.get(m.id, {})
        org_name = m.org.name if m.org else str(m.org_id)
        ws.append([
            i,
            m.hostname or "",
            m.machine_uuid[:16] + "…",
            _label_status(m.status),
            m.lifecycle or "",
            "Có" if m.is_vm else ("Không" if m.is_vm is False else ""),
            _spec_field(spec, "os_name"),
            _spec_field(spec, "os_build"),
            _cpu_label(spec),
            spec.ram_gb if spec else None,
            _disks_label(spec),
            assigned.full_name if assigned else (tmeta.get("full_name") or ""),
            assigned.email if assigned else (tmeta.get("email") or ""),
            phone_plain or "",
            tmeta.get("department") or "",
            org_name,
            _fmt_dt(m.last_seen_at),
            _fmt_dt(m.enrolled_at),
            m.note or "",
        ])

    _auto_col_width(ws, len(headers), ws.max_row)

    # ── Sheet 2: Thống kê ──────────────────────────────────────
    ws2 = wb.create_sheet("Thống kê")
    ws2["A1"] = "Thống kê tổng quan — " + datetime.now(UTC).strftime("%Y-%m-%d %H:%M")
    ws2["A1"].font = TITLE_FONT
    ws2["A2"] = f"Tổng số máy: {len(machines)}"
    ws2["A3"] = "Phân bổ theo trạng thái:"
    status_counts: dict[str, int] = {}
    for m in machines:
        status_counts[m.status] = status_counts.get(m.status, 0) + 1
    row = 4
    for status, cnt in sorted(status_counts.items()):
        ws2.cell(row=row, column=1, value=f"  • {_label_status(status)}: {cnt}")
        row += 1
    row += 1
    ws2.cell(row=row, column=1, value="Phân bổ theo tổ chức:").font = Font(bold=True)
    org_counts: dict[str, int] = {}
    for m in machines:
        name = m.org.name if m.org else str(m.org_id)
        org_counts[name] = org_counts.get(name, 0) + 1
    for name, cnt in sorted(org_counts.items(), key=lambda kv: -kv[1]):
        row += 1
        ws2.cell(row=row, column=1, value=f"  • {name}: {cnt}")
    ws2.column_dimensions["A"].width = 50

    # ── Sheet 3: Minh bạch dữ liệu (mục 7.4) ───────────────────
    ws3 = wb.create_sheet("Thông tin thu thập")
    ws3["A1"] = "THÔNG BÁO — DỮ LIỆU THU THẬP BỞI HỆ THỐNG QUẢN LÝ TÀI SẢN MÁY TÍNH"
    ws3["A1"].font = TITLE_FONT
    lines = [
        "",
        "Mục đích thu thập: quản lý tài sản máy tính (định danh, trạng thái online/offline,",
        "cấu hình, người sử dụng) phục vụ công tác quản trị tại đơn vị.",
        "",
        "Dữ liệu thu thập (chỉ đọc, không điều khiển máy):",
        "  • Cấu hình máy: OS, CPU, RAM, ổ cứng, GPU, mainboard, BIOS",
        "  • Mạng: hostname, IP, MAC (mục đích phát hiện dual-homed)", 
        "  • Phần mềm đã cài, trạng thái Antivirus / Windows Update",
        "  • User đang đăng nhập, trạng thái online/offline, thời gian bật máy",
        "",
        "KHÔNG thu thập: nội dung liên lạc, lịch sử web, phím gõ, ảnh màn hình.",
        "",
        "Tuân thủ: Nghị định 13/2023/NĐ-CP, Luật An toàn thông tin mạng 2015,",
        "Luật An ninh mạng 2018. Số điện thoại (nếu có) được mã hóa AES-256-GCM",
        "và chỉ hiển thị đầy đủ cho người có thẩm quyền.",
    ]
    for idx, line in enumerate(lines, start=2):
        cell = ws3.cell(row=idx, column=1, value=line)
        cell.alignment = WRAP
    ws3.column_dimensions["A"].width = 90

    # Xuất file
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Helpers ─────────────────────────────────────────────────────

def _latest_spec(m) -> Any | None:
    if not m.specs:
        return None
    return m.specs[0]  # relationship order_by: collected_at desc


def _spec_field(spec: Any | None, field: str) -> str:
    return str(getattr(spec, field, "") or "") if spec else ""


def _cpu_label(spec: Any | None) -> str:
    if not spec or not spec.cpu:
        return ""
    cpu = spec.cpu
    model = cpu.get("model", "")
    cores = cpu.get("cores")
    return f"{model} ({cores} cores)" if cores else str(model)


def _disks_label(spec: Any | None) -> str:
    if not spec or not spec.disks:
        return ""
    parts = []
    for d in spec.disks:
        model = d.get("model", "")
        cap = d.get("capacity_gb")
        parts.append(f"{model} {cap}GB".strip())
    return "; ".join(parts) if parts else ""


def _label_status(status: str) -> str:
    labels = {
        "online": "Online",
        "offline": "Offline",
        "lost": "Mất liên lạc (máy ma)",
        "decommissioned": "Đã thanh lý",
        "pending": "Chờ duyệt",
    }
    return labels.get(status, status)


def _decrypt_phone(assigned) -> str:
    from app.core.security import decrypt_aes_gcm

    try:
        return decrypt_aes_gcm(assigned.phone_encrypted)
    except Exception:  # noqa: BLE001
        return ""


# ── PDF (Phase 4 — WeasyPrint từ HTML template) ───────────────

_PDF_HTML = """<!doctype html>
<html lang="vi">
<head>
<meta charset="utf-8">
<style>
  @page {{ size: A4 landscape; margin: 14mm; }}
  body {{ font-family: "Segoe UI", Arial, sans-serif; font-size: 10px; color: #1e293b; }}
  h1 {{ font-size: 17px; margin: 0 0 2px; }}
  .meta {{ color: #64748b; font-size: 10px; margin-bottom: 10px; }}
  table {{ border-collapse: collapse; width: 100%; }}
  th {{ background: #1f4e78; color: #fff; padding: 5px 6px; text-align: left; }}
  td {{ border: 1px solid #cbd5e1; padding: 4px 6px; }}
  tr:nth-child(even) td {{ background: #f8fafc; }}
  .notice {{ margin-top: 12px; font-size: 8.5px; color: #64748b; }}
</style>
</head>
<body>
<h1>DANH SÁCH MÁY TÍNH — HỆ THỐNG QUẢN LÝ TÀI SẢN</h1>
<p class="meta">Tổng số: {total} máy · Xuất lúc {generated} · Người xuất: {generated_by}</p>
<table>
<thead><tr><th>STT</th><th>Hostname</th><th>Mã máy</th><th>Trạng thái</th><th>Vòng đời</th>
<th>OS</th><th>CPU</th><th>RAM</th><th>Người dùng</th><th>Điện thoại</th><th>Tổ chức</th><th>Lần cuối online</th></tr></thead>
<tbody>{rows}</tbody>
</table>
<p class="notice">Dữ liệu thu thập phục vụ quản lý tài sản — tuân thủ Nghị định 13/2023/NĐ-CP.
Số điện thoại được mã hóa AES-256-GCM; hiển thị đầy đủ chỉ cho người có thẩm quyền.</p>
</body>
</html>"""


def build_machines_pdf(
    machines: list[Any],
    *,
    token_meta: dict[uuid.UUID, dict] | None = None,
    show_full_phone: bool = False,
    generated_by: str | None = None,
) -> bytes:
    """Báo cáo PDF theo biểu mẫu hành chính (WeasyPrint, Phase 4)."""
    from weasyprint import HTML

    token_meta = token_meta or {}
    rows = []
    for i, m in enumerate(machines, start=1):
        spec = _latest_spec(m)
        assigned = m.assigned_user
        phone_plain = ""
        if m.assigned_user_id and hasattr(assigned, "phone_encrypted"):
            phone_plain = (
                mask_phone(assigned.phone_encrypted)
                if not show_full_phone
                else _decrypt_phone(assigned)
            )
        tmeta = token_meta.get(m.id, {})
        user_name = assigned.full_name if assigned else (tmeta.get("full_name") or "")
        rows.append(
            "<tr>"
            f"<td>{i}</td>"
            f"<td>{m.hostname or ''}</td>"
            f"<td>{m.machine_uuid[:12]}…</td>"
            f"<td>{_label_status(m.status)}</td>"
            f"<td>{m.lifecycle or ''}</td>"
            f"<td>{_spec_field(spec, 'os_name')} {_spec_field(spec, 'os_build')}</td>"
            f"<td>{_cpu_label(spec)}</td>"
            f"<td>{spec.ram_gb if spec and spec.ram_gb else ''}</td>"
            f"<td>{user_name}</td>"
            f"<td>{phone_plain}</td>"
            f"<td>{m.org.name if m.org else str(m.org_id)[:8]}</td>"
            f"<td>{_fmt_dt(m.last_seen_at)}</td>"
            "</tr>"
        )
    html = _PDF_HTML.format(
        total=len(machines),
        generated=datetime.now(UTC).strftime("%Y-%m-%d %H:%M"),
        generated_by=generated_by or "",
        rows="\n".join(rows),
    )
    return HTML(string=html).write_pdf()